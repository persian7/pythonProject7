# copy sheet a sheet from desk to destination
import openpyxl as xl
import openpyxl
import pandas as pd
from openpyxl import load_workbook

source = r"\\ACCURATEMATE\projects\Test Slips\Sep 28.xlsx"
destination = r"\\ACCURATEMATE\office-folder\Back up\9-23\9-1.xlsx"

SS = pd.read_excel(source, None)
S = list(SS.keys())

dd = pd.read_excel(destination, None)
b = list(dd.keys())

wb = load_workbook(destination)
for i in range(1, len(b)):
    if b[i] in wb.sheetnames:
        wb.remove(wb[b[i]])
        print(b[i])
    wb.save(destination)

from pandas import ExcelWriter

path1 = source
# destination path:
path2 = destination
wb = openpyxl.load_workbook(path1)
# give the full path of the file here
sheetname = (wb.sheetnames)
for item in range(0,(len(sheetname)), 1):
    print(sheetname[item])
    wb1 = xl.load_workbook(filename=path1)
    ws1 = wb1.worksheets[item]

    wb2 = xl.load_workbook(filename=path2)
    ws2 = wb2.create_sheet(ws1.title)

    for row in ws1:
        for cell in row:
            ws2[cell.coordinate].value = cell.value

    wb2.save(path2)

print(len(sheetname))
