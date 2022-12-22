# copy sheet a sheet from desk to destination
import openpyxl as xl
import openpyxl

path1 = r"\\ACCURATEMATE\projects\Test Slips\Sep 28.xlsx"
# destination path:
path2 = r"\\ACCURATEMATE\office-folder\Back up\9-23\9-1.xlsx"
wb = openpyxl.load_workbook(path1)
# give the full path of the file here
sheetname = (wb.sheetnames)
for item in range(0,(len(sheetname)), 1):
    print(sheetname[item])
    wb1 = xl.load_workbook(filename=path1)
    ws1 = wb1.worksheets[item]

    wb2 = xl.load_workbook(filename=path2)
    ws2 = wb2.create_sheet(ws1.title)

    for row in ws1:
        for cell in row:
            ws2[cell.coordinate].value = cell.value

    wb2.save(path2)

print(len(sheetname))
