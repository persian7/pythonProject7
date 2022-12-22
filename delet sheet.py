from openpyxl import load_workbook
import pandas as pd

source = r"\\ACCURATEMATE\projects\Test Slips\Sep 19.xlsx"
destination = r"C:\Users\Office\Desktop\pp.xlsx"
dd = pd.read_excel(destination, None)
b = list(dd.keys())

for i in range(1, len(b)):
    wb = load_workbook(destination)
    if b[i] in wb.sheetnames:
        wb.remove(wb[b[i]])
        print(b[i])
    wb.save(destination)
